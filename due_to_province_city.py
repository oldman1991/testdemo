# -*- coding: UTF-8 -*-
import cProfile
import json
import pstats

from openpyxl import load_workbook


def do_work():
    wb = load_workbook(filename=u'province_city.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    result = {}
    for rx in range(2, ws.max_row + 1):
        province = ws.cell(row=rx, column=1).value
        citys = [ws.cell(row=rx, column=col).value for col in range(2, ws.max_column) if
                 ws.cell(row=rx, column=col).value]
        print(province)
        if province:
            result[province] = citys
    print(result)
    with open('province_city.json', 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False)


cProfile.run('do_work()', 'restats')
p = pstats.Stats('restats')
p.strip_dirs().sort_stats('time').print_stats()